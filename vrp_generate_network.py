# coding=gb18030

'''
activity based space_time_network code
'''
import xlrd
import numpy
import pandas as pd    
import openpyxl
from xlrd import open_workbook
from xlutils.copy import copy

MAX_LABEL_COST = 1000000000

g_train_node_list = []
g_train_link_list = []
g_passenger_node_list = []
g_passenger_link_list = []
g_passenger_list = []
g_train_list = []

g_number_of_train_nodes = 0
g_number_of_train_links = 0
g_number_of_passenger_nodes = 0
g_number_of_passenger_links = 0
g_number_of_passengers = 0
g_number_of_trains = 0

g_number_of_time_intervals = 20 + 1


class Train_Node:
    def __init__(self):
        self.node_id = 0
        self.ingoing_node_list = []
        self.ingoing_link_list = []
        self.outgoing_node_list = []
        self.outgoing_link_list = []


class Train_Link:
    def __init__(self):
        self.link_id = 0
        self.from_node_id = 0
        self.to_node_id = 0
        self.type = 0
        self.cost = 0
        self.travel_time = 0
        self.coupled_passenger_link_id = -1
        self.time_dependent_LR_multiplier = []
        self.time_dependent_ADMM_multiplier = []
       
        
class Train:
    def __init__(self):
        self.train_id = 0
        self.from_node_id = 0
        self.to_node_id = 0
        self.old_from_node_id = 0
        self.old_to_node_id = 0
        self.departure_time = 0
        self.arrival_time = 0
        self.volume = 0
        self.node_sequence = []
        self.time_sequence = []
        self.link_sequence = []
        self.available_node_list = []
        self.time_dependent_link_volume = []
        self.time_dependent_link_binary = []
        self.time_dependent_LR_multiplier = []
        self.time_dependent_pick_up_cost = []
        self.available_passenger_list = []
        

class Passenger_Node:
    def __init__(self):
        self.node_id = 0
        self.ingoing_node_list = []
        self.ingoing_link_list = []
        self.outgoing_node_list = []
        self.outgoing_link_list = []


class Passenger_Link:
    def __init__(self):
        self.link_id = 0
        self.from_node_id = 0
        self.to_node_id = 0
        self.type = 0
        self.cost = 0
        self.travel_time = 0
        self.coupled_train_link_id = -1
        self.time_dependent_link_cost = []



class Passenger:
    def __init__(self):
        self.passenger_id = 0
        self.from_node_id = 0
        self.to_node_id = 0
        self.old_from_node_id = 0
        self.old_to_node_id = 0
        self.departure_time = 0
        self.departure_time_window = 0
        self.arrival_time = 0
        self.arrival_time_window = 0
        self.volume = 0
        self.initial_price = 0
        self.travel_time_budget = 0
        self.actual_travel_time = 0
        self.lower_bound_travel_time = 0
        self.node_sequence = []
        self.time_sequence = []
        self.link_sequence = []
        self.node_sequence_upper_bound = []
        self.time_sequence_upper_bound = []
        self.link_sequence_upper_bound = []
        self.available_node_list = []
        self.time_dependent_LR_multiplier = []
        self.time_dependent_link_volume = []
        self.time_dependent_link_binary = []

        self.time_dependent_link_volume_upper_bound = []



def g_read_input_data():    
    #initialization
    global g_number_of_train_nodes
    global g_number_of_train_links
    global g_number_of_trains
    global g_number_of_passenger_nodes
    global g_number_of_passenger_links
    global g_number_of_passengers

    # read train_node
    file = xlrd.open_workbook("input_train_node.xlsx") # open file
    sheet = file.sheet_by_index(0) # open sheet
    train_node = Train_Node()
    train_node.node_id = 0
    g_train_node_list.append(train_node)
    g_number_of_train_nodes += 1    
    for row in range(1, sheet.nrows):
        try:
            train_node = Train_Node()
            train_node.node_id = int(sheet.cell_value(row,0))
            g_train_node_list.append(train_node)
            g_number_of_train_nodes += 1
            print('reading {} train nodes..'.format(g_number_of_train_nodes))
        except:
            print('Read error. Check your train node file')
    print('train nodes_number:{}'.format(g_number_of_train_nodes))  
    
    # read train_links
    file = xlrd.open_workbook("input_train_link.xlsx") # open file
    sheet = file.sheet_by_index(0) # open sheet
    train_link = Train_Link()
    train_link.link_id = 0
    g_train_link_list.append(train_link)
    g_number_of_train_links += 1
    for row in range(1, sheet.nrows):
        try:
            train_link = Train_Link()
            train_link.link_id = int(sheet.cell_value(row, 0))
            train_link.from_node_id = int(sheet.cell_value(row, 1))
            train_link.to_node_id = int(sheet.cell_value(row, 2))
            train_link.type = int(sheet.cell_value(row, 3))
            train_link.cost = int(sheet.cell_value(row, 4))
            train_link.travel_time = int(sheet.cell_value(row, 5))
            train_link.coupled_passenger_link_id = int(sheet.cell_value(row, 6))
            g_train_link_list.append(train_link)
            g_number_of_train_links += 1
            print('reading {} train links..'.format(g_number_of_train_links))
        except:
            print('Read error. Check your train link file')
    print('train links_number:{}'.format(g_number_of_train_links))
    
    # read trains
    file = xlrd.open_workbook("input_train.xlsx") # open file
    sheet = file.sheet_by_index(0) # open sheet
    train = Train()
    train.train_id = 0
    g_train_list.append(train)
    g_number_of_trains += 1
    for row in range(1, sheet.nrows):
        try:
            train = Train()
            train.train_id = int(sheet.cell_value(row, 0))
            train.from_node_id = int(sheet.cell_value(row, 1))
            train.to_node_id = int(sheet.cell_value(row, 2))
            train.departure_time = int(sheet.cell_value(row, 3))
            train.arrival_time = int(sheet.cell_value(row, 4))
            train.volume = int(sheet.cell_value(row, 5))
            available_node_list = str(sheet.cell_value(row, 6))
            train.available_node_list = available_node_list.strip().split(';')
            # transfer str to int
            train.available_node_list = [int(node) for node in train.available_node_list]
            train.available_passenger_list = int(sheet.cell_value(row, 7))
            g_train_list.append(train)
            g_number_of_trains += 1
            print('reading {} trains..'.format(g_number_of_trains))
        except:
            print('Read. Check your train file')
    print('trains_number:{}'.format(g_number_of_trains))  
    
    
    # read passenger_node
    file = xlrd.open_workbook("input_passenger_node.xlsx") # open file
    sheet = file.sheet_by_index(0) # open sheet
    passenger_node = Passenger_Node()
    passenger_node.node_id = 0
    g_passenger_node_list.append(passenger_node)
    g_number_of_passenger_nodes += 1    
    for row in range(1, sheet.nrows):
        try:
            passenger_node = Passenger_Node()
            passenger_node.node_id = int(sheet.cell_value(row,0))
            g_passenger_node_list.append(passenger_node)
            g_number_of_passenger_nodes += 1
            print('reading {} passenger nodes..'.format(g_number_of_passenger_nodes))
        except:
            print('Read error. Check your passenger node file')
    print('passenger nodes_number:{}'.format(g_number_of_passenger_nodes))  
    
    # read passenger_links
    file = xlrd.open_workbook("input_passenger_link.xlsx") # open file
    sheet = file.sheet_by_index(0) # open sheet
    passenger_link = Passenger_Link()
    passenger_link.link_id = 0
    g_passenger_link_list.append(passenger_link)
    g_number_of_passenger_links += 1
    for row in range(1, sheet.nrows):
        try:
            passenger_link = Passenger_Link()
            passenger_link.link_id = int(sheet.cell_value(row, 0))
            passenger_link.from_node_id = int(sheet.cell_value(row, 1))
            passenger_link.to_node_id = int(sheet.cell_value(row, 2))
            passenger_link.type = int(sheet.cell_value(row, 3))
            passenger_link.cost = int(sheet.cell_value(row, 4))
            passenger_link.travel_time = int(sheet.cell_value(row, 5))
            passenger_link.coupled_train_link_id = int(sheet.cell_value(row, 6))
            g_passenger_link_list.append(passenger_link)
            g_number_of_passenger_links += 1
            print('reading {} passenger links..'.format(g_number_of_passenger_links))
        except:
            print('Read error. Check your passenger link file')
    print('passenger links_number:{}'.format(g_number_of_passenger_links))
    
    # read passengers
    file = xlrd.open_workbook("input_passenger.xlsx") # open file
    sheet = file.sheet_by_index(0) # open sheet
    passenger = Passenger()
    passenger.passenger_id = 0
    g_passenger_list.append(passenger)
    g_number_of_passengers += 1
    for row in range(1, sheet.nrows):
        try:
            passenger = Passenger()
            passenger.passenger_id = int(sheet.cell_value(row, 0))
            passenger.from_node_id = int(sheet.cell_value(row, 1))
            passenger.to_node_id = int(sheet.cell_value(row, 2))
            passenger.departure_time = int(sheet.cell_value(row, 3))
            passenger.arrival_time = int(sheet.cell_value(row, 4))
            passenger.volume = int(sheet.cell_value(row, 5))
            available_node_list = str(sheet.cell_value(row, 6))
            passenger.available_node_list = available_node_list.strip().split(';')
            # transfer str to int
            passenger.available_node_list = [int(node) for node in passenger.available_node_list]
            passenger.initial_price = int(sheet.cell_value(row, 7))
            passenger.travel_time_budget = int(sheet.cell_value(row, 8))
            passenger.departure_time_window = int(sheet.cell_value(row, 9))
            passenger.arrival_time_window = int(sheet.cell_value(row, 10))    # arrival time window[arrival_time-window,arrival_time]
            g_passenger_list.append(passenger)
            g_number_of_passengers += 1
            print('reading {} passengers..'.format(g_number_of_passengers))
        except:
            print('Read. Check your passenger file')
    print('passengers_number:{}'.format(g_number_of_passengers))

def g_add_new_node():
    global g_number_of_passenger_nodes
    global g_number_of_passengers
    global g_number_of_train_nodes
    global g_number_of_trains
    for p in range(1, g_number_of_passengers):
        new_node_pickup = Passenger_Node()
        new_node_pickup.node_id = int(500+g_passenger_list[p].passenger_id)
        g_passenger_node_list.append(new_node_pickup)
        g_passenger_list[p].old_from_node_id=g_passenger_list[p].from_node_id
        g_passenger_list[p].from_node_id=new_node_pickup.node_id
        g_passenger_list[p].available_node_list.append(new_node_pickup.node_id)
        g_number_of_passenger_nodes += 1
        g_train_node_list.append(new_node_pickup)
        g_number_of_train_nodes += 1
        for i in range(1,g_number_of_trains):
            g_train_list[i].available_node_list.append(new_node_pickup.node_id)
        # ===========Ð´Èëpassenger node input=============================
        rexcel = open_workbook("input_passenger_node.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_node_pickup.node_id
        row = rows
        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        excel.save("input_passenger_node.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel

        rexcel = open_workbook("input_train_node.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_node_pickup.node_id
        row = rows
        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        excel.save("input_train_node.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel

#====================ÐéÄâ½ÓÈËµã================================
        new_node_delivery = Passenger_Node()
        new_node_delivery.node_id = int(600+g_passenger_list[p].passenger_id)
        g_passenger_node_list.append(new_node_delivery)
        g_passenger_list[p].old_to_node_id = g_passenger_list[p].to_node_id
        g_passenger_list[p].to_node_id = new_node_delivery.node_id
        g_number_of_passenger_nodes += 1
        g_passenger_list[p].available_node_list.append(new_node_delivery.node_id)
        g_train_node_list.append(new_node_delivery)
        g_number_of_train_nodes += 1
        for i in range(1, g_number_of_trains):
            g_train_list[i].available_node_list.append(new_node_delivery.node_id)

        # ===========Ð´Èëpassenger node input=============================
        rexcel = open_workbook("input_passenger_node.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_node_delivery.node_id
        row = rows
        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        excel.save("input_passenger_node.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel

        rexcel = open_workbook("input_train_node.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_node_delivery.node_id
        row = rows
        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        excel.save("input_train_node.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel
#==========================³µÁ¾³öÈë¿â=============================
    for k in range(1, g_number_of_trains):
        new_node_out = Train_Node()
        new_node_out.node_id = int(700+g_train_list[k].train_id)
        g_train_list[k].old_from_node_id = g_train_list[k].from_node_id
        g_train_list[k].from_node_id= new_node_out.node_id
        g_train_list[k].available_node_list.append(new_node_out.node_id)
        g_number_of_train_nodes += 1
        g_train_node_list.append(new_node_out)
        # ===========Ð´Èëtrain node input=============================
        rexcel = open_workbook("input_train_node.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_node_out.node_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ

        excel.save("input_train_node.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel

        new_node_in =Train_Node()
        new_node_in.node_id = int(800+g_train_list[k].train_id)
        g_train_list[k].old_to_node_id = g_train_list[k].to_node_id
        g_train_list[k].to_node_id = new_node_in.node_id
        g_train_list[k].available_node_list.append(new_node_in.node_id)
        g_train_node_list.append(new_node_in)
        g_number_of_train_nodes += 1

        # ===========Ð´Èëtrain node input=============================
        rexcel = open_workbook("input_train_node.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_node_in.node_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ

        excel.save("input_train_node.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel

def g_add_new_passenger_link(): #passenger pick up link for both train and passenger network
    global g_number_of_passenger_nodes
    global g_number_of_passenger_links
    global g_number_of_train_links
    global g_number_of_passengers
    for p in range(1, g_number_of_passengers):
        new_link_pickup = Passenger_Link()
        new_link_pickup.link_id = 0
        new_link_delivery = Passenger_Link()
        new_link_delivery.link_id = 0
        from_node_id = g_passenger_list[p].old_from_node_id
        new_link_pickup.link_id = int(g_passenger_list[p].passenger_id)*1000+int(from_node_id)
        new_link_pickup.from_node_id = int(500+g_passenger_list[p].passenger_id)
        new_link_pickup.to_node_id = from_node_id
        new_link_pickup.type = 100
        new_link_pickup.cost = 0
        new_link_pickup.travel_time = 1
        new_link_pickup.coupled_train_link_id = new_link_pickup.link_id
        g_passenger_link_list.append(new_link_pickup)



        #===========Ð´Èëpassenger pickup=============================
        rexcel = open_workbook("input_passenger_link.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_link_pickup.link_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        table.write(row, 1, new_link_pickup.from_node_id)
        table.write(row, 2, new_link_pickup.to_node_id)
        table.write(row, 3, new_link_pickup.type)
        table.write(row, 4, new_link_pickup.cost)
        table.write(row, 5, new_link_pickup.travel_time)

        table.write(row, 6, new_link_pickup.coupled_train_link_id)
        excel.save("input_passenger_link.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel
        g_number_of_passenger_links += 1


        # ===========Ð´Èëtrain pickup=============================
        new_link_pickup = Train_Link()
        from_node_id = g_passenger_list[p].old_from_node_id
        new_link_pickup.link_id = int(g_passenger_list[p].passenger_id) * 1000 + int(from_node_id)
        new_link_pickup.from_node_id = int(500 + g_passenger_list[p].passenger_id)
        new_link_pickup.to_node_id = from_node_id
        new_link_pickup.type = 100
        new_link_pickup.cost = 0
        new_link_pickup.travel_time = 1
        new_link_pickup.coupled_passenger_link_id =  new_link_pickup.link_id
        g_train_link_list.append(new_link_pickup)

        rexcel = open_workbook("input_train_link.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_link_pickup.link_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        table.write(row, 1, new_link_pickup.from_node_id)
        table.write(row, 2, new_link_pickup.to_node_id)
        table.write(row, 3, new_link_pickup.type)
        table.write(row, 4, new_link_pickup.cost)
        table.write(row, 5, new_link_pickup.travel_time)
        table.write(row, 6, new_link_pickup.coupled_passenger_link_id)
        excel.save("input_train_link.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel
        g_number_of_train_links += 1

       #================·µ=====================================================================
        new_link_pickup = Passenger_Link()
        new_link_pickup.link_id = int(g_passenger_list[p].passenger_id) * 1000 + 10 + int(from_node_id)
        new_link_pickup.to_node_id = int(500 + g_passenger_list[p].passenger_id)
        new_link_pickup.from_node_id = from_node_id
        new_link_pickup.type = 100
        new_link_pickup.cost = 0
        new_link_pickup.travel_time = 1
        new_link_pickup.coupled_train_link_id = new_link_pickup.link_id
        g_passenger_link_list.append(new_link_pickup)

        rexcel = open_workbook("input_passenger_link.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_link_pickup.link_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        table.write(row, 1, new_link_pickup.from_node_id)
        table.write(row, 2, new_link_pickup.to_node_id)
        table.write(row, 3, new_link_pickup.type)
        table.write(row, 4, new_link_pickup.cost)
        table.write(row, 5, new_link_pickup.travel_time)
        table.write(row, 6, values)
        excel.save("input_passenger_link.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel
        g_number_of_passenger_links += 1

        new_link_pickup = Train_Link()
        new_link_pickup.link_id = int(g_passenger_list[p].passenger_id) * 1000 + 10 + int(from_node_id)
        new_link_pickup.to_node_id = int(500 + g_passenger_list[p].passenger_id)
        new_link_pickup.from_node_id = from_node_id
        new_link_pickup.type = 100
        new_link_pickup.cost = 0
        new_link_pickup.travel_time = 1
        new_link_pickup.coupled_passenger_link_id = new_link_pickup.link_id
        g_train_link_list.append(new_link_pickup)

        rexcel = open_workbook("input_train_link.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_link_pickup.link_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        table.write(row, 1, new_link_pickup.from_node_id)
        table.write(row, 2, new_link_pickup.to_node_id)
        table.write(row, 3, new_link_pickup.type)
        table.write(row, 4, new_link_pickup.cost)
        table.write(row, 5, new_link_pickup.travel_time)
        table.write(row, 6, new_link_pickup.coupled_passenger_link_id)
        excel.save("input_train_link.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel
        g_number_of_train_links += 1

        #==============µÈ´ý====================
        new_link_pickup= Passenger_Link()
        new_link_pickup.link_id = int(g_passenger_list[p].passenger_id) * 1000 + 50 + int(from_node_id)
        new_link_pickup.to_node_id = int(500 + g_passenger_list[p].passenger_id)
        new_link_pickup.from_node_id = int(500 + g_passenger_list[p].passenger_id)
        new_link_pickup.type = 2
        new_link_pickup.cost = 0
        new_link_pickup.travel_time = 1
        new_link_pickup.coupled_train_link_id = new_link_pickup.link_id
        g_passenger_link_list.append(new_link_pickup)

        rexcel = open_workbook("input_passenger_link.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_link_pickup.link_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        table.write(row, 1, new_link_pickup.from_node_id)
        table.write(row, 2, new_link_pickup.to_node_id)
        table.write(row, 3, new_link_pickup.type)
        table.write(row, 4, new_link_pickup.cost)
        table.write(row, 5, new_link_pickup.travel_time)
        table.write(row, 6, values)
        excel.save("input_passenger_link.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel
        g_number_of_passenger_links += 1

        new_link_pickup = Train_Link()
        new_link_pickup.link_id = int(g_passenger_list[p].passenger_id) * 1000 + 50 + int(from_node_id)
        new_link_pickup.to_node_id = int(500 + g_passenger_list[p].passenger_id)
        new_link_pickup.from_node_id = int(500 + g_passenger_list[p].passenger_id)
        new_link_pickup.type = 2
        new_link_pickup.cost = 0
        new_link_pickup.travel_time = 1
        new_link_pickup.coupled_passenger_link_id = new_link_pickup.link_id
        g_train_link_list.append(new_link_pickup)
        rexcel = open_workbook("input_train_link.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_link_pickup.link_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        table.write(row, 1, new_link_pickup.from_node_id)
        table.write(row, 2, new_link_pickup.to_node_id)
        table.write(row, 3, new_link_pickup.type)
        table.write(row, 4, new_link_pickup.cost)
        table.write(row, 5, new_link_pickup.travel_time)
        table.write(row, 6, new_link_pickup.coupled_passenger_link_id)
        excel.save("input_train_link.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel
        g_number_of_train_links += 1

        ###   ====================ËÍÈËlink=========================
        to_node_id = g_passenger_list[p].old_to_node_id
        new_link_delivery.link_id = int(g_passenger_list[p].passenger_id) * 1000 + int(to_node_id)
        new_link_delivery.from_node_id = to_node_id
        new_link_delivery.to_node_id = int(600+g_passenger_list[p].passenger_id)
        new_link_delivery.type = 200
        new_link_delivery.cost = 0
        new_link_delivery.travel_time = 1
        new_link_delivery.coupled_train_link_id = new_link_delivery.link_id
        g_passenger_link_list.append(new_link_delivery)
        g_number_of_passenger_links += 1
        # ===========Ð´Èëpassenger link input=============================
        rexcel = open_workbook("input_passenger_link.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_link_delivery.link_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        table.write(row, 1, new_link_delivery.from_node_id)
        table.write(row, 2, new_link_delivery.to_node_id)
        table.write(row, 3, new_link_delivery.type)
        table.write(row, 4, new_link_delivery.cost)
        table.write(row, 5, new_link_delivery.travel_time)
        table.write(row, 6, new_link_delivery.coupled_train_link_id)
        excel.save("input_passenger_link.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel

        # ===========Ð´Èëtrain link input=============================
        new_link_delivery = Train_Link()
        new_link_delivery.link_id = int(g_passenger_list[p].passenger_id) * 1000 + int(to_node_id)
        new_link_delivery.from_node_id = to_node_id
        new_link_delivery.to_node_id = int(600 + g_passenger_list[p].passenger_id)
        new_link_delivery.type = 200
        new_link_delivery.cost = 0
        new_link_delivery.travel_time = 1
        new_link_delivery.coupled_passenger_link_id = new_link_delivery.link_id
        g_train_link_list.append(new_link_delivery)

        rexcel = open_workbook("input_train_link.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_link_delivery.link_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        table.write(row, 1, new_link_delivery.from_node_id)
        table.write(row, 2, new_link_delivery.to_node_id)
        table.write(row, 3, new_link_delivery.type)
        table.write(row, 4, new_link_delivery.cost)
        table.write(row, 5, new_link_delivery.travel_time)
        table.write(row, 6, new_link_delivery.coupled_passenger_link_id)
        excel.save("input_train_link.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel
        g_number_of_train_links += 1

        # ================·µ=====================================================================
        new_link_delivery = Passenger_Link()
        new_link_delivery.link_id = int(g_passenger_list[p].passenger_id) * 1000 + 10 + int(to_node_id)
        new_link_delivery.from_node_id = int(600 + g_passenger_list[p].passenger_id)
        new_link_delivery.to_node_id = to_node_id
        new_link_delivery.type = 200
        new_link_delivery.cost = 0
        new_link_delivery.travel_time = 1
        new_link_delivery.coupled_train_link_id = new_link_delivery.link_id
        g_passenger_link_list.append(new_link_delivery)

        rexcel = open_workbook("input_passenger_link.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_link_delivery.link_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        table.write(row, 1, new_link_delivery.from_node_id)
        table.write(row, 2, new_link_delivery.to_node_id)
        table.write(row, 3, new_link_delivery.type)
        table.write(row, 4, new_link_delivery.cost)
        table.write(row, 5, new_link_delivery.travel_time)
        table.write(row, 6, values)
        excel.save("input_passenger_link.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel
        g_number_of_passenger_links += 1
        #==============================================

        new_link_delivery = Train_Link()
        new_link_delivery.link_id = int(g_passenger_list[p].passenger_id) * 1000 + 10 + int(to_node_id)
        new_link_delivery.to_node_id = to_node_id
        new_link_delivery.from_node_id = int(600 + g_passenger_list[p].passenger_id)
        new_link_delivery.type = 200
        new_link_delivery.cost = 0
        new_link_delivery.travel_time = 1
        new_link_delivery.coupled_passenger_link_id = new_link_delivery.link_id
        g_train_link_list.append(new_link_delivery)

        rexcel = open_workbook("input_train_link.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_link_delivery.link_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        table.write(row, 1, new_link_delivery.from_node_id)
        table.write(row, 2, new_link_delivery.to_node_id)
        table.write(row, 3, new_link_delivery.type)
        table.write(row, 4, new_link_delivery.cost)
        table.write(row, 5, new_link_delivery.travel_time)
        table.write(row, 6, new_link_delivery.coupled_passenger_link_id)
        excel.save("input_train_link.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel
        g_number_of_train_links += 1

        # ==============µÈ´ý====================
        new_link_delivery = Passenger_Link()
        new_link_delivery.link_id = int(g_passenger_list[p].passenger_id) * 1000 + 50 + int(to_node_id)
        new_link_delivery.to_node_id = int(600 + g_passenger_list[p].passenger_id)
        new_link_delivery.from_node_id = int(600 + g_passenger_list[p].passenger_id)
        new_link_delivery.type = 2
        new_link_delivery.cost = 0
        new_link_delivery.travel_time = 1
        new_link_delivery.coupled_train_link_id = new_link_delivery.link_id
        g_passenger_link_list.append(new_link_delivery)

        rexcel = open_workbook("input_passenger_link.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_link_delivery.link_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        table.write(row, 1, new_link_delivery.from_node_id)
        table.write(row, 2, new_link_delivery.to_node_id)
        table.write(row, 3, new_link_delivery.type)
        table.write(row, 4, new_link_delivery.cost)
        table.write(row, 5, new_link_delivery.travel_time)
        table.write(row, 6, values)
        excel.save("input_passenger_link.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel
        g_number_of_passenger_links += 1

        new_link_delivery = Train_Link()
        new_link_delivery.link_id = int(g_passenger_list[p].passenger_id) * 1000 + 50 + int(to_node_id)
        new_link_delivery.to_node_id = int(600 + g_passenger_list[p].passenger_id)
        new_link_delivery.from_node_id = int(600 + g_passenger_list[p].passenger_id)
        new_link_delivery.type = 2
        new_link_delivery.cost = 0
        new_link_delivery.travel_time = 1
        new_link_delivery.coupled_passenger_link_id = new_link_delivery.link_id
        g_train_link_list.append(new_link_delivery)
        rexcel = open_workbook("input_train_link.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_link_delivery.link_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        table.write(row, 1, new_link_delivery.from_node_id)
        table.write(row, 2, new_link_delivery.to_node_id)
        table.write(row, 3, new_link_delivery.type)
        table.write(row, 4, new_link_delivery.cost)
        table.write(row, 5, new_link_delivery.travel_time)
        table.write(row, 6, new_link_delivery.coupled_passenger_link_id)
        excel.save("input_train_link.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel
        g_number_of_train_links += 1




    return()

def g_add_new_train_link():
    global g_number_of_train_nodes
    global g_number_of_train_links
    global g_number_of_trains


    for k in range(1, g_number_of_trains):
        new_link_out = Train_Link()
        new_link_out.link_id = 0
        new_link_in =Train_Link()
        new_link_in.link_id = 0
        from_node_id = g_train_list[k].old_from_node_id
        new_link_out.link_id = int(g_train_list[k].train_id) * 10000 + int(from_node_id)
        new_link_out.from_node_id = int(700+g_train_list[k].train_id)
        new_link_out.to_node_id = from_node_id
        new_link_out.type = 1000
        new_link_out.travel_time = 1
        new_link_out.coupled_passenger_link_id = new_link_out.link_id
        new_link_out.cost = 0
        g_train_link_list.append(new_link_out)
        g_number_of_train_links += 1
        #=====================new link out==============type 2000 ±íÊ¾³µÈë¿â 1000±íÊ¾³µ³ö¿â==
        rexcel = open_workbook("input_train_link.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_link_out.link_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        table.write(row, 1, new_link_out.from_node_id)
        table.write(row, 2, new_link_out.to_node_id)
        table.write(row, 3, new_link_out.type)
        table.write(row, 4, new_link_out.cost)
        table.write(row, 5, new_link_out.travel_time)
        table.write(row, 6, new_link_out.coupled_passenger_link_id)
        excel.save("input_train_link.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel

        new_link_out = Train_Link()
        new_link_out.link_id = int(g_train_list[k].train_id) * 10000 + 10 + int(from_node_id)
        new_link_out.to_node_id = int(700 + g_train_list[k].train_id)
        new_link_out.from_node_id = from_node_id
        new_link_out.type = 1000
        new_link_out.travel_time = 1
        new_link_out.coupled_passenger_link_id = new_link_out.link_id
        new_link_out.cost = 0
        g_train_link_list.append(new_link_out)
        g_number_of_train_links += 1

        rexcel = open_workbook("input_train_link.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_link_out.link_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        table.write(row, 1, new_link_out.from_node_id)
        table.write(row, 2, new_link_out.to_node_id)
        table.write(row, 3, new_link_out.type)
        table.write(row, 4, new_link_out.cost)
        table.write(row, 5, new_link_out.travel_time)
        table.write(row, 6, new_link_out.coupled_passenger_link_id)
        excel.save("input_train_link.xlsx")

        new_link_out = Train_Link()
        new_link_out.link_id = int(g_train_list[k].train_id) * 10000 + 50 + int(from_node_id)
        new_link_out.to_node_id = int(700 + g_train_list[k].train_id)
        new_link_out.from_node_id = int(700 + g_train_list[k].train_id)
        new_link_out.type = 2
        new_link_out.travel_time = 1
        new_link_out.coupled_passenger_link_id = new_link_out.link_id
        new_link_out.cost = 0
        g_train_link_list.append(new_link_out)
        g_number_of_train_links += 1

        rexcel = open_workbook("input_train_link.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_link_out.link_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        table.write(row, 1, new_link_out.from_node_id)
        table.write(row, 2, new_link_out.to_node_id)
        table.write(row, 3, new_link_out.type)
        table.write(row, 4, new_link_out.cost)
        table.write(row, 5, new_link_out.travel_time)
        table.write(row, 6, new_link_out.coupled_passenger_link_id)
        excel.save("input_train_link.xlsx")



        #=====================new link in ================
        new_link_in = Train_Link()
        to_node_id = g_train_list[k].old_to_node_id
        new_link_in.link_id = int(g_train_list[k].train_id) * 10000 + int(to_node_id)
        new_link_in.from_node_id = to_node_id
        new_link_in.to_node_id = int(800+g_train_list[k].train_id)
        new_link_in.type = 2000
        new_link_in.travel_time = 1
        new_link_in.coupled_passenger_link_id =  new_link_in.link_id
        new_link_in.cost = 0
        g_train_link_list.append(new_link_in)
        g_number_of_train_links += 1
        # ===========Ð´Èënew link in=============================
        rexcel = open_workbook("input_train_link.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_link_in.link_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        table.write(row, 1, new_link_in.from_node_id)
        table.write(row, 2, new_link_in.to_node_id)
        table.write(row, 3, new_link_in.type)
        table.write(row, 4, new_link_in.cost)
        table.write(row, 5, new_link_in.travel_time)
        table.write(row, 6, new_link_in.coupled_passenger_link_id)
        excel.save("input_train_link.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel

        new_link_in = Train_Link()
        new_link_in.link_id = int(g_train_list[k].train_id) * 10000 + 10 + int(to_node_id)
        new_link_in.to_node_id = to_node_id
        new_link_in.from_node_id = int(800 + g_train_list[k].train_id)
        new_link_in.type = 2000
        new_link_in.travel_time = 1
        new_link_in.coupled_passenger_link_id = new_link_in.link_id
        new_link_in.cost = 0
        g_train_link_list.append(new_link_in)
        g_number_of_train_links += 1
        # ===========Ð´Èënew link in=============================
        rexcel = open_workbook("input_train_link.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_link_in.link_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        table.write(row, 1, new_link_in.from_node_id)
        table.write(row, 2, new_link_in.to_node_id)
        table.write(row, 3, new_link_in.type)
        table.write(row, 4, new_link_in.cost)
        table.write(row, 5, new_link_in.travel_time)
        table.write(row, 6, new_link_in.coupled_passenger_link_id)
        excel.save("input_train_link.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel


        new_link_in = Train_Link()
        new_link_in.link_id = int(g_train_list[k].train_id) * 10000 + 50 + int(to_node_id)
        new_link_in.to_node_id = int(800 + g_train_list[k].train_id)
        new_link_in.from_node_id = int(800 + g_train_list[k].train_id)
        new_link_in.type = 2
        new_link_in.travel_time = 1
        new_link_in.coupled_passenger_link_id = new_link_in.link_id
        new_link_in.cost = 0
        g_train_link_list.append(new_link_in)
        g_number_of_train_links += 1
        # ===========Ð´Èënew link in=============================
        rexcel = open_workbook("input_train_link.xlsx")  # ÓÃwlrdÌá¹©µÄ·½·¨¶ÁÈ¡Ò»¸öexcelÎÄ¼þ
        rows = rexcel.sheets()[0].nrows  # ÓÃwlrdÌá¹©µÄ·½·¨»ñµÃÏÖÔÚÒÑÓÐµÄÐÐÊý
        excel = copy(rexcel)  # ÓÃxlutilsÌá¹©µÄcopy·½·¨½«xlrdµÄ¶ÔÏó×ª»¯ÎªxlwtµÄ¶ÔÏó
        table = excel.get_sheet(0)  # ÓÃxlwt¶ÔÏóµÄ·½·¨»ñµÃÒª²Ù×÷µÄsheet
        values = new_link_in.link_id
        row = rows

        table.write(row, 0, values)  # xlwt¶ÔÏóµÄÐ´·½·¨£¬²ÎÊý·Ö±ðÊÇÐÐ¡¢ÁÐ¡¢Öµ
        table.write(row, 1, new_link_in.from_node_id)
        table.write(row, 2, new_link_in.to_node_id)
        table.write(row, 3, new_link_in.type)
        table.write(row, 4, new_link_in.cost)
        table.write(row, 5, new_link_in.travel_time)
        table.write(row, 6, new_link_in.coupled_passenger_link_id)
        excel.save("input_train_link.xlsx")  # xlwt¶ÔÏóµÄ±£´æ·½·¨£¬ÕâÊ±±ã¸²¸ÇµôÁËÔ­À´µÄexcel

def  g_generate_in_out_going_link():
    # record ingoing and outgoing nodes and links for each train node
    for l in range(1, g_number_of_train_links):
        link_id = g_train_link_list[l].link_id
        from_node_id = g_train_link_list[l].from_node_id
        to_node_id = g_train_link_list[l].to_node_id
        for dd in range(1, g_number_of_train_nodes):
            if g_train_node_list[dd].node_id == from_node_id:
                from_node_id_index = dd
                for ll in range(1,g_number_of_train_nodes):
                    if g_train_node_list[ll].node_id == to_node_id:
                        to_node_id_index = ll
                        g_train_node_list[to_node_id_index].ingoing_link_list.append(link_id)
                        g_train_node_list[to_node_id_index].ingoing_node_list.append(from_node_id)
                        g_train_node_list[from_node_id_index].outgoing_link_list.append(link_id)
                        g_train_node_list[from_node_id_index].outgoing_node_list.append(to_node_id)




    # record ingoing and outgoing nodes and links for each passenger node
    for l in range(1, g_number_of_passenger_links):
        link_id = g_passenger_link_list[l].link_id
        from_node_id = g_passenger_link_list[l].from_node_id
        to_node_id = g_passenger_link_list[l].to_node_id
        for dd in range(1, g_number_of_passenger_nodes):
            if g_passenger_node_list[dd].node_id == from_node_id:
                from_node_id_index = dd
                for ll in range(1,g_number_of_passenger_nodes):
                    if g_passenger_node_list[ll].node_id == to_node_id:
                        to_node_id_index = ll
                        g_passenger_node_list[to_node_id_index].ingoing_link_list.append(link_id)
                        g_passenger_node_list[to_node_id_index].ingoing_node_list.append(from_node_id)
                        g_passenger_node_list[from_node_id_index].outgoing_link_list.append(link_id)
                        g_passenger_node_list[from_node_id_index].outgoing_node_list.append(to_node_id)
    return()

def g_generate_node_link_map():
    # from node to node map to link id
    global train_node_link_map
    train_node_link_map = pd.DataFrame({'link_id':[],'from_node_id':[],'to_node_id':[]})    
    for l in range(1, g_number_of_train_links):
        train_node_link_map = train_node_link_map.append({'link_id':g_train_link_list[l].link_id,'from_node_id':g_train_link_list[l].from_node_id,'to_node_id':g_train_link_list[l].to_node_id},ignore_index=True)
        
    # from node to node map to link id
    global passenger_node_link_map
    passenger_node_link_map = pd.DataFrame({'link_id':[],'from_node_id':[],'to_node_id':[]})    
    for l in range(1, g_number_of_passenger_links):
        passenger_node_link_map = passenger_node_link_map.append({'link_id':g_passenger_link_list[l].link_id,'from_node_id':g_passenger_link_list[l].from_node_id,'to_node_id':g_passenger_link_list[l].to_node_id},ignore_index=True)
        
    return()

def g_time_dependent_dynamic_programming_for_passengers():

     #initial time-dependnt-link-volume
    for p in range (1, g_number_of_passengers):
        g_passenger_list[p].time_dependent_link_volume = [[0 for t in range(1, g_number_of_time_intervals + 1)] for l in range(1, g_number_of_passenger_links + 1)]
        g_passenger_list[p].time_dependent_link_binary = [[0 for t in range(1, g_number_of_time_intervals + 1)] for l in range(1, g_number_of_passenger_links + 1)]

        g_passenger_list[p].lower_bound_travel_time = 0
    # initialize time dependent link cost
    time_dependent_link_cost = -100000 * numpy.ones([g_number_of_passenger_links, g_number_of_time_intervals])
    # dynamic programming for each passenger
    for p in range(1, g_number_of_passengers):         
        pre_node_id = - numpy.ones([g_number_of_passenger_nodes, g_number_of_time_intervals])
        pre_time_interval = - numpy.ones([g_number_of_passenger_nodes, g_number_of_time_intervals])
        label_cost = - MAX_LABEL_COST * numpy.ones([g_number_of_passenger_nodes, g_number_of_time_intervals])
        g_passenger_list[p].node_sequence = []
        g_passenger_list[p].time_sequence= []   
        g_passenger_list[p].link_sequence= []  
        # get agent information
        from_node_id = int (g_passenger_list[p].from_node_id)
        to_node_id = int (g_passenger_list[p].to_node_id)
        departure_time = int (g_passenger_list[p].departure_time)
        travel_time_budget = int (g_passenger_list[p].travel_time_budget)
        dummy_arrival_time = min(departure_time + travel_time_budget, g_number_of_time_intervals - 1)
        volume = g_passenger_list[p].volume
        # set time dependent link cost

        for l in range(1, g_number_of_passenger_links):
            for t in range(1, g_number_of_time_intervals):
                cost = g_passenger_link_list[l].cost
                time_dependent_link_cost[l][t] = cost
        # initialize

        # dynamic programming

        for j in range(1,g_number_of_passenger_nodes):
            if g_passenger_node_list[j].node_id==from_node_id:
                from_node_id_index=j
                pre_node_id[from_node_id_index][departure_time] = 0
                pre_time_interval[from_node_id_index][departure_time] = departure_time
                label_cost[from_node_id_index][departure_time] = 0

        for t in range(departure_time, dummy_arrival_time):
            for n in g_passenger_list[p].available_node_list:
                for nn in range(1,g_number_of_passenger_nodes):
                    if g_passenger_node_list[nn].node_id==n:
                        n_index=nn
                        if (pre_node_id[n_index][t] != -1):
                            for l in range(0, len(g_passenger_node_list[n_index].outgoing_link_list)):
                                outgoing_link_id = g_passenger_node_list[n_index].outgoing_link_list[l]
                                outgoing_node_id = g_passenger_node_list[n_index].outgoing_node_list[l]
                                for i in range(0, len(g_passenger_link_list)):
                                    if g_passenger_link_list[i].link_id == outgoing_link_id:
                                        outgoing_link_id_index = i
                                        if (g_passenger_link_list[outgoing_link_id_index].coupled_train_link_id == -1):
                                            # cost = time_dependent_link_cost[outgoing_link_id_index][t]
                                            cost = g_passenger_link_list[outgoing_link_id_index].time_dependent_link_cost[t]
                                            trave_time = g_passenger_link_list[outgoing_link_id_index].travel_time
                                            if (t + trave_time < g_number_of_time_intervals):
                                                for kk in range(0, len(g_passenger_node_list)):
                                                    if g_passenger_node_list[kk].node_id == outgoing_node_id:
                                                        outgoing_node_id_index = kk

                                                        if (label_cost[n_index][t] + cost >= label_cost[outgoing_node_id_index][t + trave_time]):

                                                            label_cost[outgoing_node_id_index][t + trave_time] = label_cost[n_index][t] + cost
                                                            pre_node_id[outgoing_node_id_index][t + trave_time] = n
                                                            pre_time_interval[outgoing_node_id_index][t + trave_time] = t
                                        if (g_passenger_link_list[outgoing_link_id_index].coupled_train_link_id != -1):
                                            coupled_train_link_id = g_passenger_link_list[outgoing_link_id_index].coupled_train_link_id

                                            for w in range(1, g_number_of_train_links):
                                                if g_train_link_list[w].link_id == coupled_train_link_id:
                                                    coupled_train_link_id_index = w
                                                    cost = g_passenger_link_list[outgoing_link_id_index].time_dependent_link_cost[t] + \
                                                           g_passenger_list[p].time_dependent_LR_multiplier[outgoing_link_id_index][t] + g_train_link_list[coupled_train_link_id_index].time_dependent_LR_multiplier[t]  ##¸üÐÂcost
                                                    trave_time = g_passenger_link_list[outgoing_link_id_index].travel_time
                            # cost = cost - g_train_link_list[outgoing_link_id].time_dependent_ADMM_multiplier[t]
                                                    if (t + trave_time < g_number_of_time_intervals):
                                                        for kk in range(0, len(g_passenger_node_list)):
                                                            if g_passenger_node_list[kk].node_id == outgoing_node_id:
                                                                outgoing_node_id_index = kk
                                                                if (label_cost[n_index][t] + cost >= label_cost[outgoing_node_id_index][t + trave_time]):
                                                                    label_cost[outgoing_node_id_index][t + trave_time] = label_cost[n_index][t] + cost
                                                                    pre_node_id[outgoing_node_id_index][t + trave_time] = n
                                                                    pre_time_interval[outgoing_node_id_index][t + trave_time] = t



        # backtrace
        n = to_node_id
        t = dummy_arrival_time
        g_passenger_list[p].node_sequence.insert(0, n)
        g_passenger_list[p].time_sequence.insert(0, t)
        for w in range(1,g_number_of_passenger_nodes):
            if g_passenger_node_list[w].node_id==n:
                to_node_id_index=w
                if (label_cost[to_node_id_index][t] == MAX_LABEL_COST):
                    print('can not find space-time path for passenger:{}'.format(p))
                for backtrace_step in range(1, g_number_of_time_intervals):
                    if (label_cost[to_node_id_index][t] != MAX_LABEL_COST):
                        if (n != from_node_id) or (t != departure_time):
                            pre_n = int (pre_node_id[to_node_id_index][t])
                            pre_t = int (pre_time_interval[to_node_id_index][t])
                    # get current link id
                            if (pre_n!= -1) and (pre_t!= -1):
                                l = int (passenger_node_link_map.loc[(passenger_node_link_map.from_node_id == pre_n) & (passenger_node_link_map.to_node_id == n),'link_id'].values[0])
                    # update time-dependent link volume
                                for ll in range(1,g_number_of_passenger_links):
                                    if g_passenger_link_list[ll].link_id==l:
                                        l_index=ll
                                        time_dependent_link_volume_for_passengers[l_index][pre_t] += volume
                                        g_passenger_list[p].time_dependent_link_volume[l_index][pre_t] += volume
                                        g_passenger_list[p].time_dependent_link_binary[l_index][pre_t] = int(g_passenger_list[p].time_dependent_link_volume[l_index][pre_t] > 0)
                                n = pre_n
                                for w in range(1, g_number_of_passenger_nodes):
                                    if g_passenger_node_list[w].node_id == n:
                                        to_node_id_index = w
                                t = pre_t
                    # node and time sequence lists
                                g_passenger_list[p].node_sequence.insert(0, n)
                                g_passenger_list[p].time_sequence.insert(0, t)
                                g_passenger_list[p].link_sequence.insert(0, l)
                                        # print(g_passenger_list[k].node_sequence)
        #       calcualte actual travel time
                g_passenger_list[p].actual_travel_time = 0
                for l in range(1, g_number_of_passenger_links):
                    for t in range(1, g_number_of_time_intervals):
                        g_passenger_list[p].lower_bound_travel_time = g_passenger_list[p].lower_bound_travel_time + g_passenger_list[p].time_dependent_link_volume[l][t] * g_passenger_link_list[l].cost

    return ()
        #=====================================================
    #     for t in range(departure_time, dummy_arrival_time):
    #         for n in g_passenger_list[p].available_node_list:
    #             for nn in range(1,g_number_of_passenger_nodes):
    #                 if g_passenger_node_list[nn].node_id==n:
    #                     n_index=nn
    #                     if (pre_node_id[n_index][t] != -1):            # only choose the departure node
    #                         for l in range(0, len(g_passenger_node_list[n_index].outgoing_link_list)):     #¶ÔÓÚÆðµãµÄ¿ÉÐÐ»¡
    #                             outgoing_link_id = g_passenger_node_list[n_index].outgoing_link_list[l]
    #                             outgoing_node_id = g_passenger_node_list[n_index].outgoing_node_list[l]
    #                             for i in range(0,len(g_passenger_link_list)):
    #                                 if g_passenger_link_list[i].link_id==outgoing_link_id:
    #                                     outgoing_link_id_index=i
    #                                     if (g_passenger_link_list[outgoing_link_id_index].coupled_train_link_id == -1): #Ö»ÄÜÈË×ßµÄ»¡
    #                                         cost = g_passenger_link_list[outgoing_link_id_index].time_dependent_link_cost[t]
    #                                         trave_time = g_passenger_link_list[outgoing_link_id_index].travel_time
    #                                         if (t + trave_time < g_number_of_time_intervals):
    #                                             # for nn in range(1,g_number_of_passenger_nodes):
    #                                             #     if g_passenger_node_list[nn].node_id==n:
    #                                             #         n_index=nn
    #                                                     for kk in range(1,g_number_of_passenger_nodes):
    #                                                         if g_passenger_node_list[kk].node_id==outgoing_node_id:
    #                                                             outgoing_node_id_index=kk
    #                                                             if (label_cost[n_index][t] + cost >= label_cost[outgoing_node_id_index][t + trave_time]):
    #                                                                 label_cost[outgoing_node_id_index][t + trave_time] = label_cost[ n_index][t] + cost
    #                                                                 pre_node_id[outgoing_node_id_index][t + trave_time] = n
    #                                                                 pre_time_interval[outgoing_node_id_index][t + trave_time] = t
    #                                     if (g_passenger_link_list[outgoing_link_id_index].coupled_train_link_id != -1):
    #                                         coupled_train_link_id = g_passenger_link_list[outgoing_link_id_index].coupled_train_link_id
    #                                         for w in range(1,g_number_of_train_links):
    #                                             if g_train_link_list[w].link_id== coupled_train_link_id:
    #                                                 coupled_train_link_id_index=w
    #
    #                                                 cost = g_passenger_link_list[outgoing_link_id_index].time_dependent_link_cost[t] + g_passenger_list[p].time_dependent_LR_multiplier[outgoing_link_id_index][t] + g_train_link_list[coupled_train_link_id_index].time_dependent_LR_multiplier[t]   ##¸üÐÂcost
    #                                                 trave_time = g_passenger_link_list[outgoing_link_id_index].travel_time
    #                                                 if (t + trave_time < g_number_of_time_intervals):
    #                                                     for j in range(0, len(g_passenger_node_list)):
    #                                                         if g_passenger_node_list[j].node_id == outgoing_node_id:
    #                                                             outgoing_node_id_index = j
    #                                                             for nn in range(1, g_number_of_passenger_nodes):
    #                                                                 if g_passenger_node_list[nn].node_id == n:
    #                                                                     n_index = nn
    #                                                                     if (label_cost[n_index][t] + cost >= label_cost[outgoing_node_id_index][t + trave_time]):
    #                                                                         label_cost[outgoing_node_id_index][t + trave_time] = label_cost[n_index][t] + cost
    #                                                                         pre_node_id[outgoing_node_id_index][t + trave_time] = n
    #                                                                         pre_time_interval[outgoing_node_id_index][t + trave_time] = t
    #     # backtrace
    #     n = to_node_id
    #
    #     t = dummy_arrival_time
    #     g_passenger_list[p].node_sequence.insert(0, n)            #ÔÚµÚ0¸öÎ»ÖÃ²åÈën
    #     g_passenger_list[p].time_sequence.insert(0, t)
    #     for d in range (1,g_number_of_passenger_nodes):
    #         if g_passenger_node_list[d].node_id==n:
    #             to_node_id_index=d
    #             if (label_cost[to_node_id_index][t] == MAX_LABEL_COST):
    #                 print('can not find space-time path for passenger:{}'.format(p))
    #             for backtrace_step in range(1, g_number_of_time_intervals):
    #                 if (label_cost[to_node_id_index][t] != MAX_LABEL_COST):
    #                     if (n != from_node_id) or (t != departure_time):
    #                         pre_n = int (pre_node_id[to_node_id_index][t])
    #                         pre_t = int (pre_time_interval[to_node_id_index][t])
    #                 # get current link id
    #                         l = int (passenger_node_link_map.loc[(passenger_node_link_map.from_node_id == pre_n) & (passenger_node_link_map.to_node_id == n),'link_id'].values[0])
    #                 # update time-dependent link volume
    #                         for ll in range(1,g_number_of_passenger_links):
    #                             if g_passenger_link_list[ll].link_id==l:
    #                                 l_index=ll
    #                                 time_dependent_link_volume_for_passengers[l_index][pre_t] += volume
    #                                 g_passenger_list[p].time_dependent_link_volume[l_index][pre_t] += volume
    #                                 g_passenger_list[p].time_dependent_link_binary[l_index][pre_t]=int(g_passenger_list[p].time_dependent_link_volume[l_index][pre_t] >0)
    #                                 n = pre_n
    #                                 t = pre_t
    #                 # node and time sequence lists
    #                                 g_passenger_list[p].node_sequence.insert(0, n)
    #                                 g_passenger_list[p].time_sequence.insert(0, t)
    #                                 g_passenger_list[p].link_sequence.insert(0, l)
    #     # calcualte actual travel time
    #             g_passenger_list[p].actual_travel_time = 0
    #             for l in range(1, g_number_of_passenger_links):
    #                 for t in range(1, g_number_of_time_intervals):
    #                     g_passenger_list[p].lower_bound_travel_time = g_passenger_list[p].lower_bound_travel_time + g_passenger_list[p].time_dependent_link_volume[l][t] * g_passenger_link_list[l].cost
    #
    # return()

def g_time_dependent_dynamic_programming_for_trains():
    global time_dependent_link_cost_for_pick_up

    #initial time dependnt link volume
    for k in range (1, g_number_of_trains):
        g_train_list[k].time_dependent_link_volume = [[0 for t in range(1, g_number_of_time_intervals + 1)] for l in range(1, g_number_of_train_links + 1)]
        g_train_list[k].time_dependent_link_binary = [[0 for t in range(1, g_number_of_time_intervals + 1)] for l in range(1, g_number_of_train_links + 1)]
    # initialize time dependent link cost
        time_dependent_link_cost = - 100000 * numpy.ones([g_number_of_train_links, g_number_of_time_intervals])
    # dynamic programming for each train
    for k in range(1, g_number_of_trains):          
        pre_node_id = - numpy.ones([g_number_of_train_nodes, g_number_of_time_intervals])
        pre_time_interval = - numpy.ones([g_number_of_train_nodes, g_number_of_time_intervals])
        label_cost = - MAX_LABEL_COST * numpy.ones([g_number_of_train_nodes, g_number_of_time_intervals])
        g_train_list[k].node_sequence = []
        g_train_list[k].time_sequence= []   
        g_train_list[k].link_sequence= []  
        # get agent information
        from_node_id = int (g_train_list[k].from_node_id)
        to_node_id = int (g_train_list[k].to_node_id)
        departure_time = int (g_train_list[k].departure_time)
        arrival_time = int (g_train_list[k].arrival_time)
        volume = g_train_list[k].volume
        # set time dependent link cost
        for l in range(1, g_number_of_train_links):
            for t in range(1, g_number_of_time_intervals):
                cost = g_train_link_list[l].cost
                time_dependent_link_cost[l][t] = cost
        # set time dependent link cost
# =============================================================================
#         for l in range(1, g_number_of_train_links):
#             for t in range(1, g_number_of_time_intervals):
#                 cost = g_train_link_list[l].cost
#                 time_dependent_link_cost[l][t] = cost
#                 if l == 21:
#                     for t in range(5, 6):
#                         time_dependent_link_cost[l][t] = 10
#                 # if l == 20:
#                 #     for t in range(6, 7):
#                 #         time_dependent_link_cost[l][t] = 2
# 
#                 if l == 27:
#                     if t in range(8, 9):
#                         time_dependent_link_cost[l][t] = 10
#                 # if l== 26:
#                 #     if t in range(9,10):
#                 #         time_dependent_link_cost[l][t] = 2
# 
#                 if l == 23:
#                     for t in range(10, 11):
#                         time_dependent_link_cost[l][t] = 10
# 
#                 # if l == 24:
#                 #     for t in range(11, 12):
#                 #         time_dependent_link_cost[l][t] = 2
# 
#                 if l == 29:
# 
#                     for t in range(13, 14):
#                         time_dependent_link_cost[l][t] = 10
# =============================================================================

                # if l == 30:
                #
                #     for t in range(14, 15):
                #         time_dependent_link_cost[l][t] = 2

                # if l==8:
                #     for t in range(7,8):
                #         time_dependent_link_cost[l][t] = 3
                #     # for t in range(10, 12):
                #     #     time_dependent_link_cost[l][t] = 3
                # if l == 12:
                #     for t in range(9, 10):
                #         time_dependent_link_cost[l][t] = 10
                #     for t in range(12, 15):
                #         time_dependent_link_cost[l][t] = 3
                # if l == 11:
                #     for t in range(9, 11):
                #         time_dependent_link_cost[l][t] = 1

        # initialize
        for j in range(1,g_number_of_train_nodes):
            if g_train_node_list[j].node_id==from_node_id:
                from_node_id_index=j
                pre_node_id[from_node_id_index][departure_time] = 0
                pre_time_interval[from_node_id_index][departure_time] = departure_time
                label_cost[from_node_id_index][departure_time] = 0
        # dynamic programming
        for t in range(departure_time, arrival_time):
            for n in g_train_list[k].available_node_list:
                for nn in range(1,g_number_of_train_nodes):
                    if g_train_node_list[nn].node_id == n:
                        n_index=nn
                        if (pre_node_id[n_index][t] != -1):
                            for l in range(0, len(g_train_node_list[n_index].outgoing_link_list)):
                                outgoing_link_id = g_train_node_list[n_index].outgoing_link_list[l]
                                outgoing_node_id = g_train_node_list[n_index].outgoing_node_list[l]
                                for i in range(0, len(g_train_link_list)):
                                    if g_train_link_list[i].link_id == outgoing_link_id:
                                        outgoing_link_id_index = i
                                        if (g_train_link_list[outgoing_link_id_index].coupled_passenger_link_id == -1):
                                            cost = g_train_list[k].time_dependent_link_cost_for_pick_up[outgoing_link_id_index][t]

                                            trave_time = g_train_link_list[outgoing_link_id_index].travel_time
                                            if (t + trave_time < g_number_of_time_intervals):
                                                for kk in range(0, len(g_train_node_list)):
                                                    if g_train_node_list[kk].node_id == outgoing_node_id:
                                                        outgoing_node_id_index = kk

                                                        if (label_cost[n_index][t] + cost >= label_cost[outgoing_node_id_index][t + trave_time]):

                                                            label_cost[outgoing_node_id_index][t + trave_time] = label_cost[n_index][t] + cost
                                                            pre_node_id[outgoing_node_id_index][t + trave_time] = n
                                                            pre_time_interval[outgoing_node_id_index][t + trave_time] = t
                                        if (g_train_link_list[outgoing_link_id_index].coupled_passenger_link_id != -1):
                                            coupled_passenger_link_id = g_train_link_list[outgoing_link_id_index].coupled_passenger_link_id
                                            cost = g_train_list[k].time_dependent_link_cost_for_pick_up[outgoing_link_id_index][t]

                                            for p in range(1, g_number_of_passengers):
                                                for pp in range(1,g_number_of_passenger_links):
                                                    if g_passenger_link_list[pp].link_id==coupled_passenger_link_id:
                                                        coupled_passenger_link_id_index=pp
                                                        cost = cost + g_passenger_list[p].time_dependent_LR_multiplier[coupled_passenger_link_id_index][t]
                            # cost = cost - g_train_link_list[outgoing_link_id].time_dependent_ADMM_multiplier[t]
                                            trave_time = g_train_link_list[outgoing_link_id_index].travel_time
                                            if (t + trave_time < g_number_of_time_intervals):
                                                for kk in range(0, len(g_train_node_list)):
                                                    if g_train_node_list[kk].node_id == outgoing_node_id:
                                                        outgoing_node_id_index = kk
                                                        if (label_cost[n_index][t] + cost >= label_cost[outgoing_node_id_index][t + trave_time]):

                                                            label_cost[outgoing_node_id_index][t + trave_time] = label_cost[n_index][t] + cost
                                                            pre_node_id[outgoing_node_id_index][t + trave_time] = n
                                                            pre_time_interval[outgoing_node_id_index][t + trave_time] = t



        # backtrace
        n = to_node_id
        t = arrival_time
        g_train_list[k].node_sequence.insert(0, n)
        g_train_list[k].time_sequence.insert(0, t)
        for w in range(1,g_number_of_train_nodes):
            if g_train_node_list[w].node_id==n:
                to_node_id_index=w
                if (label_cost[to_node_id_index][t] == MAX_LABEL_COST):
                    print('can not find space-time path for train:{}'.format(k))
                for backtrace_step in range(1, g_number_of_time_intervals):
                    if (label_cost[to_node_id_index][t] != MAX_LABEL_COST):
                        if (n != from_node_id) or (t != departure_time):
                            pre_n = int (pre_node_id[to_node_id_index][t])
                            pre_t = int (pre_time_interval[to_node_id_index][t])
                    # get current link id
                            if (pre_n!= -1) and (pre_t!= -1):
                                l = int (train_node_link_map.loc[(train_node_link_map.from_node_id == pre_n) & (train_node_link_map.to_node_id == n),'link_id'].values[0])
                    # update time-dependent link volume
                                for ll in range(1,g_number_of_train_links):
                                    if g_train_link_list[ll].link_id==l:
                                        l_index=ll
                                        time_dependent_link_volume_for_trains[l_index][pre_t] += volume
                                        g_train_list[k].time_dependent_link_volume[l_index][pre_t] += volume
                                        g_train_list[k].time_dependent_link_binary[l_index][pre_t] = int(g_train_list[k].time_dependent_link_volume[l_index][pre_t] > 0)
                                n = pre_n
                                for w in range(1, g_number_of_train_nodes):
                                    if g_train_node_list[w].node_id == n:
                                        to_node_id_index = w
                                t = pre_t
                    # node and time sequence lists
                                g_train_list[k].node_sequence.insert(0, n)
                                g_train_list[k].time_sequence.insert(0, t)
                                g_train_list[k].link_sequence.insert(0, l)
                                print(g_train_list[k].node_sequence)
# =============================================================================
#         # ====É¾³ý³µ1·þÎñµã
#         count_first_train=0
#         count_second_train=0
# 
#         if k < 2:
#             if 501 in g_train_list[k].available_node_list and 601 in g_train_list[k].available_node_list:
#                 for pre_t in range(5, 6):
#                     for l in range(1,g_number_of_train_links):
#                         if l == 21:
#                             if g_train_list[k].time_dependent_link_binary[l][pre_t] == 1:
#                                 count_first_train  = 1
#                 for pre_t in range(10, 11):
#                     for l in range(1, g_number_of_train_links):
#                         if l == 23:
#                             if g_train_list[k].time_dependent_link_binary[l][pre_t] == 1:
#                                 count_first_train += 1
#                 if count_first_train == 2:
#                     if 501 in g_train_list[k+1].available_node_list and 601 in g_train_list[k+1].available_node_list:
#                         g_train_list[k + 1].available_node_list.remove(501)
#                         g_train_list[k + 1].available_node_list.remove(601)
# 
#             if 502 in g_train_list[k].available_node_list and 602 in g_train_list[k].available_node_list:
#                 for pre_t in range(9, 11):
#                     if l == 26:
#                         if g_train_list[k].time_dependent_link_binary[l][pre_t] == 1:
#                             count_second_train = 1
#                 for pre_t in range(13, 15):
#                     if l == 29:
#                         if g_train_list[k].time_dependent_link_binary[l][pre_t] == 1:
#                             count_second_train += 1
#                 if count_second_train == 2:
#                     if 502 in g_train_list[k+1].available_node_list and 602 in g_train_list[k+1].available_node_list:
#                         g_train_list[k + 1].available_node_list.remove(502)
#                         g_train_list[k + 1].available_node_list.remove(602)
# 
#         print(numpy.nonzero(g_train_list[k].time_dependent_link_binary))
# 
#     # print(count_first_train)
#     # print(count_second_train)
# =============================================================================










        # update ADMM multiplier
    #     for l in range(1, g_number_of_train_links):
    #         for t in range(1, g_number_of_time_intervals):
    #             if (g_train_link_list[l].type == 1):
    #                 for kk in range(1, g_number_of_trains):
    #                     if (g_train_list[kk].time_dependent_link_volume[l][t] == 1):
    #                         for tt in range(max(1, t - headway), t + headway):
    #                             g_train_link_list[l].time_dependent_ADMM_multiplier[tt] = rou
    #
    # # re-assignment for train
    # # strat re-assignment
    # for re_assignment_iteration in range(1, 3):
    #     print(re_assignment_iteration)
    #     for k in range(1, g_number_of_trains):
    #         pre_node_id = - numpy.ones([g_number_of_train_nodes, g_number_of_time_intervals])
    #         pre_time_interval = - numpy.ones([g_number_of_train_nodes, g_number_of_time_intervals])
    #         label_cost = - MAX_LABEL_COST * numpy.ones([g_number_of_train_nodes, g_number_of_time_intervals])
    #         # g_train_list[k].node_sequence = []
    #         # g_train_list[k].time_sequence= []
    #         # g_train_list[k].link_sequence= []
    #         # get agent information
    #         from_node_id = int (g_train_list[k].from_node_id)
    #         to_node_id = int (g_train_list[k].to_node_id)
    #         departure_time = int (g_train_list[k].departure_time)
    #         arrival_time = int (g_train_list[k].arrival_time)
    #         volume = g_train_list[k].volume
    #         # delete the current volume
    #         for l in range(0, len(g_train_list[k].link_sequence)):
    #             link_id = g_train_list[k].link_sequence[l]
    #             time = g_train_list[k].time_sequence[l]
    #             g_train_list[k].time_dependent_link_volume[link_id][time] = g_train_list[k].time_dependent_link_volume[link_id][time] - volume
    #             time_dependent_link_volume_for_trains[link_id][time] = time_dependent_link_volume_for_trains[link_id][time] - volume
    #         # set time dependent link cost
    #         for l in range(1, g_number_of_train_links):
    #             for t in range(1, g_number_of_time_intervals):
    #                 cost = g_train_link_list[l].cost
    #                 time_dependent_link_cost[l][t] = cost
    #         # update ADMM multiplier
    #         #initial train link ADMM multipliers
    #         for l in range(1, g_number_of_train_links):
    #             g_train_link_list[l].time_dependent_ADMM_multiplier = [0 for t in range(0, g_number_of_time_intervals + 1)]
    #         for l in range(1, g_number_of_train_links):
    #             for t in range(1, g_number_of_time_intervals):
    #                 if (g_train_link_list[l].type == 1):
    #                     for kk in range(1, g_number_of_trains):
    #                         if (g_train_list[kk].time_dependent_link_volume[l][t] == 1):
    #                             for tt in range(max(1, t - headway), t + headway):
    #                                 g_train_link_list[l].time_dependent_ADMM_multiplier[tt] = rou
    #         # reset
    #         g_train_list[k].node_sequence = []
    #         g_train_list[k].time_sequence= []
    #         g_train_list[k].state_sequence= []
    #         g_train_list[k].link_sequence= []
    #         # initialize
    #         pre_node_id[from_node_id][departure_time] = 0
    #         pre_time_interval[from_node_id][departure_time] = departure_time
    #         label_cost[from_node_id][departure_time] = 0
    #         # dynamic programming
    #         for t in range(departure_time, arrival_time):
    #             for n in g_train_list[k].available_node_list:
    #                 if (pre_node_id[n][t] != -1):
    #                     for l in range(0, len(g_train_node_list[n].outgoing_link_list)):
    #                         outgoing_link_id = g_train_node_list[n].outgoing_link_list[l]
    #                         outgoing_node_id = g_train_node_list[n].outgoing_node_list[l]
    #                         if (g_train_link_list[outgoing_link_id].coupled_passenger_link_id == -1):
    #                             cost = 0
    #                             trave_time = g_train_link_list[outgoing_link_id].travel_time
    #                             if (t + trave_time < g_number_of_time_intervals):
    #                                 if (label_cost[n][t] + cost >= label_cost[outgoing_node_id][t + trave_time]):
    #                                     label_cost[outgoing_node_id][t + trave_time] = label_cost[n][t] + cost
    #                                     pre_node_id[outgoing_node_id][t + trave_time] = n
    #                                     pre_time_interval[outgoing_node_id][t + trave_time] = t
    #                         if (g_train_link_list[outgoing_link_id].coupled_passenger_link_id != -1):
    #                             coupled_passenger_link_id = g_train_link_list[outgoing_link_id].coupled_passenger_link_id
    #                             cost = 0
    #                             for p in range(1, g_number_of_passengers):
    #                                 cost = cost + g_passenger_list[p].time_dependent_LR_multiplier[coupled_passenger_link_id][t]
    #                             cost = cost - g_train_link_list[outgoing_link_id].time_dependent_ADMM_multiplier[t]
    #                             trave_time = g_train_link_list[outgoing_link_id].travel_time
    #                             if (t + trave_time < g_number_of_time_intervals):
    #                                 if (label_cost[n][t] + cost >= label_cost[outgoing_node_id][t + trave_time]):
    #                                     label_cost[outgoing_node_id][t + trave_time] = label_cost[n][t] + cost
    #                                     pre_node_id[outgoing_node_id][t + trave_time] = n
    #                                     pre_time_interval[outgoing_node_id][t + trave_time] = t
    #         # backtrace
    #         n = to_node_id
    #         t = arrival_time
    #         g_train_list[k].node_sequence.insert(0, n)
    #         g_train_list[k].time_sequence.insert(0, t)
    #         if (label_cost[n][t] == MAX_LABEL_COST):
    #             print('can not find space-time path for train:{}'.format(k))
    #         for backtrace_step in range(1, g_number_of_time_intervals):
    #             if (label_cost[n][t] != MAX_LABEL_COST):
    #                 if (n != from_node_id) or (t != departure_time):
    #                     pre_n = int (pre_node_id[n][t])
    #                     pre_t = int (pre_time_interval[n][t])
    #                     # get current link id
    #                     l = int (train_node_link_map.loc[(train_node_link_map.from_node_id == pre_n) & (train_node_link_map.to_node_id == n),'link_id'].values[0])
    #                     # update time-dependent link volume
    #                     time_dependent_link_volume_for_trains[l][pre_t] += volume
    #                     g_train_list[k].time_dependent_link_volume[l][pre_t] += volume
    #                     n = pre_n
    #                     t = pre_t
    #                     # node and time sequence lists
    #                     g_train_list[k].node_sequence.insert(0, n)
    #                     g_train_list[k].time_sequence.insert(0, t)
    #                     g_train_list[k].link_sequence.insert(0, l)
    #
    # return()
    #

def g_initialize_multiplier_and_cost():
    #initial passenger link cost
    for l in range(1, g_number_of_passenger_links):
        cost = g_passenger_link_list[l].cost
        g_passenger_link_list[l].time_dependent_link_cost = [cost for t in range(0, g_number_of_time_intervals + 1)]  
    #initial time-dependent passenger LR multiplier for each passegner
    for p in range(1, g_number_of_passengers):
        g_passenger_list[p].time_dependent_LR_multiplier = [[0 for t in range(1, g_number_of_time_intervals + 1)] for l in range(1, g_number_of_passenger_links + 1)]
        initial_price = g_passenger_list[p].initial_price
        # departure_time = g_passenger_list[p].departure_time
        # dummy_arrival_time = departure_time + g_passenger_list[p].travel_time_budget
        # for t in range(departure_time, min(dummy_arrival_time, g_number_of_time_intervals)):
            # for l in range(1, g_number_of_passenger_links):
                #g_passenger_list[p].time_dependent_LR_multiplier[l][t] =  initial_price
        g_passenger_list[p].time_dependent_LR_multiplier = [[initial_price for t in range(1, g_number_of_time_intervals + 1)] for l in range(1, g_number_of_passenger_links + 1)]
    #initial train link multipliers
    for l in range(1, g_number_of_train_links):
        g_train_link_list[l].time_dependent_LR_multiplier = [0 for t in range(0, g_number_of_time_intervals + 1)]
        g_train_link_list[l].time_dependent_ADMM_multiplier = [0 for t in range(0, g_number_of_time_intervals + 1)]

def g_update_multiplier():
    #update time-dependent passenger LR multiplier for each passegner
    for p in range(1, g_number_of_passengers):
        for l in range(1, g_number_of_passenger_links):
            for t in range(1, g_number_of_time_intervals):
                coupled_train_link_id = g_passenger_link_list[l].coupled_train_link_id
                if (coupled_train_link_id == -1):
                    g_passenger_list[p].time_dependent_LR_multiplier[l][t] = g_passenger_list[p].time_dependent_LR_multiplier[l][t]
                if (coupled_train_link_id != -1):
                    g_passenger_list[p].time_dependent_LR_multiplier[l][t] = max(0, g_passenger_list[p].time_dependent_LR_multiplier[l][t] + step_size * (g_passenger_list[p].time_dependent_link_binary[l][t] - time_dependent_link_volume_for_trains[l][t])) # pai
    #update time-dependent train LR multiplier
    for l in range(1, g_number_of_train_links):
        for t in range(1, g_number_of_time_intervals):
            coupled_passenger_link_id = g_train_link_list[l].coupled_passenger_link_id
            if (coupled_passenger_link_id == -1):
                g_train_link_list[l].time_dependent_LR_multiplier[t] = g_train_link_list[l].time_dependent_LR_multiplier[t]
            if (coupled_passenger_link_id != -1):
                # g_train_link_list[l].time_dependent_LR_multiplier[t] = max(0, g_train_link_list[l].time_dependent_LR_multiplier[t] + step_size * (time_dependent_link_volume_for_passengers[l][t] - cap))  #ÈËÁ÷Á¿-ÈÝÁ¿
                g_train_link_list[l].time_dependent_LR_multiplier[t] = max(0, g_train_link_list[l].time_dependent_LR_multiplier[t])
                

def g_update_cost():
    global time_dependent_link_cost_for_pick_up
    #判断当前循环乘客下界时空轨迹
    time_dependent_link_cost_for_pick_up = - 100000 * numpy.ones([g_number_of_train_links, g_number_of_time_intervals])
    
    for k in range(1, g_number_of_trains):
        g_train_list[k].time_dependent_link_cost_for_pick_up = - 100000 * numpy.ones([g_number_of_train_links, g_number_of_time_intervals])
    
    for k in range(1, g_number_of_trains):
        # for p in g_train_list[k].available_passenger_list:
        p = g_train_list[k].available_passenger_list
        
        for ii in range(0, len(g_passenger_list[p].link_sequence)):
            current_link_id = g_passenger_list[p].link_sequence[ii]
            current_time_id = g_passenger_list[p].time_sequence[ii]  
            
            for t in range(max(1, current_time_id - 1),current_time_id + 1):
                for i in range(1, g_number_of_passenger_links):
                    if (g_passenger_link_list[i].link_id==current_link_id):
                        current_link_id_index = i                    
                        g_train_list[k].time_dependent_link_cost_for_pick_up[current_link_id_index][t] = 100000
            
    
    
    
# =============================================================================
#     for p in range(1, g_number_of_passengers):
#         if (p == 1):
#             k = 1
#         if (p == 2):
#             k = 2
#             
#         for ii in range(0, len(g_passenger_list[p].link_sequence)):
#             current_link_id = g_passenger_list[p].link_sequence[ii]
#             current_time_id = g_passenger_list[p].time_sequence[ii]  
#             
#             for t in range(max(1, current_time_id - 1),current_time_id + 1):
#                 for i in range(1, g_number_of_passenger_links):
#                     if (g_passenger_link_list[i].link_id==current_link_id):
#                         current_link_id_index = i
#                     
#                         g_train_list[k].time_dependent_link_cost_for_pick_up[current_link_id_index][t] = 100000
# =============================================================================
            
                    
                    
    # g_train_list[coupled_train_id].time_dependent_LR_multiplier[current_link_id_index][current_time_id] += -100 
    # =============================================================================
#         for l in range(1, g_number_of_train_links):
#             for t in range(1, g_number_of_time_intervals):
#                 cost = g_train_link_list[l].cost
#                 time_dependent_link_cost[l][t] = cost
#                 if l == 21:
#                     for t in range(5, 6):
#                         time_dependent_link_cost[l][t] = 10
#                 # if l == 20:
#                 #     for t in range(6, 7):
#                 #         time_dependent_link_cost[l][t] = 2
# 
#                 if l == 27:
#                     if t in range(8, 9):
#                         time_dependent_link_cost[l][t] = 10
#                 # if l== 26:
#                 #     if t in range(9,10):
#                 #         time_dependent_link_cost[l][t] = 2
# 
#                 if l == 23:
#                     for t in range(10, 11):
#                         time_dependent_link_cost[l][t] = 10
# 
#                 # if l == 24:
#                 #     for t in range(11, 12):
#                 #         time_dependent_link_cost[l][t] = 2
# 
#                 if l == 29:
# 
#                     for t in range(13, 14):
#                         time_dependent_link_cost[l][t] = 10
# =============================================================================
            
          
def g_generate_and_calculate_upper_bound():
    global time_dependent_link_volume_for_passengers
    #initial time-dependnt-link-volume
    time_dependent_link_volume_for_passengers = numpy.zeros([g_number_of_passenger_links, g_number_of_time_intervals])
    for p in range (1, g_number_of_passengers):
        g_passenger_list[p].time_dependent_link_volume = [[0 for t in range(1, g_number_of_time_intervals + 1)] for l in range(1, g_number_of_passenger_links + 1)]
    # initialize time dependent link cost
    time_dependent_link_cost = - 100000 * numpy.ones([g_number_of_passenger_links, g_number_of_time_intervals])

    # dynamic programming for each passenger
    for p in range(1, g_number_of_passengers):         
        pre_node_id = - numpy.ones([g_number_of_passenger_nodes, g_number_of_time_intervals])
        pre_time_interval = - numpy.ones([g_number_of_passenger_nodes, g_number_of_time_intervals])
        label_cost = - MAX_LABEL_COST * numpy.ones([g_number_of_passenger_nodes, g_number_of_time_intervals])
        g_passenger_list[p].node_sequence_upper_bound = []
        g_passenger_list[p].time_sequence_upper_bound = []   
        g_passenger_list[p].link_sequence_upper_bound = []  
        # get agent information
        from_node_id = int (g_passenger_list[p].from_node_id)
        to_node_id = int (g_passenger_list[p].to_node_id)
        departure_time = int (g_passenger_list[p].departure_time)
        travel_time_budget = int (g_passenger_list[p].travel_time_budget)
        dummy_arrival_time = min(departure_time + travel_time_budget, g_number_of_time_intervals - 1)
        volume = g_passenger_list[p].volume


        # set time dependent link cost
        for l in range(1, g_number_of_passenger_links):
            for t in range(1, g_number_of_time_intervals):
                cost = g_passenger_link_list[l].cost
                time_dependent_link_cost[l][t] = cost
        # initialize
        for i in range(1,g_number_of_passenger_nodes):
            if g_passenger_node_list[i].node_id == from_node_id:
                from_node_id_index = i
                pre_node_id[from_node_id_index][departure_time] = 0
                pre_time_interval[from_node_id_index][departure_time] = departure_time
                label_cost[from_node_id_index][departure_time] = 0
        # dynamic programming
        for t in range(departure_time, dummy_arrival_time):        
            for n in g_passenger_list[p].available_node_list:
                for nn in range(1,g_number_of_passenger_nodes):
                    if g_passenger_node_list[nn].node_id==n:
                        n_index = nn
                        if (pre_node_id[n_index][t] != -1):
                            for l in range(0, len(g_passenger_node_list[n_index].outgoing_link_list)):
                                outgoing_link_id = g_passenger_node_list[n_index].outgoing_link_list[l]
                                outgoing_node_id = g_passenger_node_list[n_index].outgoing_node_list[l]
                                for w in range(0, len(g_passenger_link_list)):
                                    if g_passenger_link_list[w].link_id == outgoing_link_id:
                                        outgoing_link_id_index = w
                                        coupled_train_link_id = g_passenger_link_list[outgoing_link_id_index].coupled_train_link_id
                                        if (g_passenger_link_list[outgoing_link_id_index].coupled_train_link_id == -1):
                                            # cost = g_passenger_link_list[outgoing_link_id_index].time_dependent_link_cost[t]
                                            cost = 1
                                            trave_time = g_passenger_link_list[outgoing_link_id_index].travel_time
                                            if (t + trave_time < g_number_of_time_intervals):
                                                for ii in range(1,g_number_of_passenger_nodes):
                                                    if g_passenger_node_list[ii].node_id == outgoing_node_id:
                                                        outgoing_node_id_index=ii
                                                        if (label_cost[n_index][t] + cost >= label_cost[outgoing_node_id_index][t + trave_time]) :

                                                            label_cost[outgoing_node_id][t + trave_time] = label_cost[n_index][t] + cost
                                                            pre_node_id[outgoing_node_id][t + trave_time] = n
                                                            pre_time_interval[outgoing_node_id][t + trave_time] = t
                                        if (g_passenger_link_list[outgoing_link_id_index].coupled_train_link_id != -1):
                                            #coupled_train_link_id = g_passenger_link_list[outgoing_link_id_index].coupled_train_link_id
                                            for kk in range(1,g_number_of_train_links):
                                                if g_train_link_list[kk].link_id == coupled_train_link_id:
                                                    coupled_train_link_id_index = kk
                                                    if (time_dependent_link_volume_for_trains[coupled_train_link_id_index][t] >= 1):   #### volume
                                                        # coupled_train_link_id = g_passenger_link_list[outgoing_link_id_index].coupled_train_link_id
                                #                       cost = g_passenger_link_list[outgoing_link_id_index].time_dependent_link_cost[t]
                                                        cost = 100
                                                        trave_time = g_passenger_link_list[outgoing_link_id_index].travel_time
                                                        if (t + trave_time < g_number_of_time_intervals):
                                                            for ii in range(1, g_number_of_passenger_nodes):
                                                                if g_passenger_node_list[ii].node_id == outgoing_node_id:
                                                                    outgoing_node_id_index = ii
                                                                    if (label_cost[n_index][t] + cost >= label_cost[outgoing_node_id_index][t + trave_time]) and (time_dependent_link_volume_for_passengers[l][t] <= cap):
                                        #and (time_dependent_link_volume_for_passengers[g_passenger_node_list[n].outgoing_link_list[l]][t] <= cap - 35): ###ÈÝÁ¿<35
                                                                        label_cost[outgoing_node_id_index][t + trave_time] = label_cost[n_index][t] + cost
                                                                        pre_node_id[outgoing_node_id_index][t + trave_time] = n
                                                                        pre_time_interval[outgoing_node_id_index][t + trave_time] = t
        # backtrace                    
        n = to_node_id

        t = dummy_arrival_time
        penalty = g_number_of_time_intervals - departure_time
        g_passenger_list[p].node_sequence_upper_bound.insert(0, n)
        g_passenger_list[p].time_sequence_upper_bound.insert(0, t)
        for d in range(1, g_number_of_passenger_nodes):
            if g_passenger_node_list[d].node_id == n:
                to_node_id_index = d
                if (label_cost[to_node_id_index][t] == MAX_LABEL_COST):
                    print('can not find space-time path for passenger:{}'.format(p))
                for backtrace_step in range(1, g_number_of_time_intervals):
                    if (label_cost[to_node_id_index][t] != MAX_LABEL_COST):
                        if (n != from_node_id) or (t != departure_time):
                            pre_n = int (pre_node_id[to_node_id_index][t])
                            pre_t = int (pre_time_interval[to_node_id_index][t])
                    # get current link id
                            if (pre_n != -1) and (pre_t != -1):
                                l = int (passenger_node_link_map.loc[(passenger_node_link_map.from_node_id == pre_n) & (passenger_node_link_map.to_node_id == n),'link_id'].values[0])
                    # update time-dependent link volume
                                for ll in range(1,g_number_of_passenger_links):
                                    if g_passenger_link_list[ll].link_id==ll:
                                        l_index=ll
                                        time_dependent_link_volume_for_passengers[l_index][pre_t] += volume
                                        g_passenger_list[p].time_dependent_link_volume[l_index][pre_t] += volume
                                n = pre_n
                                for w in range(1, g_number_of_passenger_nodes):
                                    if g_passenger_node_list[w].node_id == n:
                                        to_node_id_index = w
                                t = pre_t
                    # node and time sequence lists
                                g_passenger_list[p].node_sequence_upper_bound.insert(0, n)
                                g_passenger_list[p].time_sequence_upper_bound.insert(0, t)
                                g_passenger_list[p].link_sequence_upper_bound.insert(0, l)
        # calculate upper bound

                for ii in range(1,g_number_of_passenger_nodes):
                    if g_passenger_node_list[ii].node_id == to_node_id:
                        to_node_id_index = ii
                        if (label_cost[to_node_id_index][dummy_arrival_time] == MAX_LABEL_COST):
                            upper_bound[iteration_step] = upper_bound[iteration_step] + penalty*g_passenger_list[p].volume
                        if (label_cost[to_node_id_index][dummy_arrival_time] != MAX_LABEL_COST):
                            for l in range(1, g_number_of_passenger_links):
                                for t in range(1, g_number_of_time_intervals):
                                    upper_bound[iteration_step] =  upper_bound[iteration_step] + g_passenger_list[p].time_dependent_link_volume[l][t] * g_passenger_link_list[l].cost

    return()    

def g_calcualte_lower_bound():
    
    # first term
    #for p in range(1, g_number_of_passengers):
    for p in range(1, g_number_of_passengers):
        lower_bound_1[iteration_step] = lower_bound_1[iteration_step] + g_passenger_list[p].lower_bound_travel_time
    # for l in range(1, g_number_of_passenger_links):
        # for t in range(1, g_number_of_time_intervals):                
            # lower_bound_1[iteration_step] = lower_bound_1[iteration_step] + time_dependent_link_volume_for_passengers[l][t] * g_passenger_link_list[l].cost
    # second term  #====
    for p in range(1, g_number_of_passengers):
        for l in range(1, g_number_of_passenger_links):
            coupled_train_link_id = g_passenger_link_list[l].coupled_train_link_id
            if (coupled_train_link_id != -1):
                for t in range(1, g_number_of_time_intervals):   
                    #lower_bound_2[iteration_step] = lower_bound_2[iteration_step] + g_passenger_list[p].time_dependent_LR_multiplier[l][t] * (g_passenger_list[p].time_dependent_link_volume[l][t] - time_dependent_link_volume_for_trains[coupled_train_link_id][t])
                    lower_bound_2[iteration_step] = lower_bound_2[iteration_step] +g_passenger_list[p].time_dependent_LR_multiplier[l][t] *g_passenger_list[p].time_dependent_link_volume[l][t]* (g_passenger_list[p].time_dependent_link_binary[l][t] -time_dependent_link_volume_for_trains[l][t])
    # third term

    for l in range(1, g_number_of_passenger_links):
        coupled_train_link_id = g_passenger_link_list[l].coupled_train_link_id
        if (coupled_train_link_id != -1):
            lower_bound_3[iteration_step] = lower_bound_3[iteration_step] + g_train_link_list[l].time_dependent_LR_multiplier[t] * (time_dependent_link_volume_for_passengers[l][t] - cap)
    
    lower_bound[iteration_step] = lower_bound_1[iteration_step] + lower_bound_2[iteration_step] + lower_bound_3[iteration_step]       
    return()

def g_write_output_data():
    
    # output passenger file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'output_passenger_results'    
    sheet['A1'] = 'passenger_id'
    sheet['B1'] = 'node_sequence'
    sheet['C1'] = 'time_sequence'
    sheet['D1'] = 'volume'
    sheet['E1'] = 'departure_time'
    sheet['F1'] = 'arrival_time'
    sheet['G1'] = 'total_travel_time'      
    for p in range(1, g_number_of_passengers):
        row = p + 1
        volume = g_passenger_list[p].volume
        departure_time = g_passenger_list[p].departure_time
        arrival_time = g_passenger_list[p].arrival_time
        total_travel_time = g_passenger_list[p].arrival_time - g_passenger_list[p].departure_time
        node_sequence = ";".join(str(node) for node in g_passenger_list[p].node_sequence)
        time_sequence = ";".join(str(time) for time in g_passenger_list[p].time_sequence)
        sheet.cell(row = row, column = 1, value = p)
        sheet.cell(row = row, column = 2, value = node_sequence)
        sheet.cell(row = row, column = 3, value = time_sequence)
        sheet.cell(row = row, column = 4, value = volume)
        sheet.cell(row = row, column = 5, value = departure_time)
        sheet.cell(row = row, column = 6, value = arrival_time)
        sheet.cell(row = row, column = 7, value = total_travel_time)           
    workbook.save('output_passenger_results.xlsx')
    
    # output train file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'output_train_results'    
    sheet['A1'] = 'train_id'
    sheet['B1'] = 'node_sequence'
    sheet['C1'] = 'time_sequence'
    sheet['D1'] = 'volume'
    sheet['E1'] = 'departure_time'
    sheet['F1'] = 'arrival_time'
    sheet['G1'] = 'total_travel_time'      
    for k in range(1, g_number_of_trains):
        row = k + 1
        volume = g_train_list[k].volume
        departure_time = g_train_list[k].departure_time
        arrival_time = g_train_list[k].arrival_time
        total_travel_time = g_train_list[k].arrival_time - g_train_list[k].departure_time
        node_sequence = ";".join(str(node) for node in g_train_list[k].node_sequence)
        time_sequence = ";".join(str(time) for time in g_train_list[k].time_sequence)
        sheet.cell(row = row, column = 1, value = k)
        sheet.cell(row = row, column = 2, value = node_sequence)
        sheet.cell(row = row, column = 3, value = time_sequence)
        sheet.cell(row = row, column = 4, value = volume)
        sheet.cell(row = row, column = 5, value = departure_time)
        sheet.cell(row = row, column = 6, value = arrival_time)
        sheet.cell(row = row, column = 7, value = total_travel_time)           
    workbook.save('output_train_results.xlsx') 
    
    # output passenger file upper bound
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'output_upper_bound'    
    sheet['A1'] = 'passenger_id'
    sheet['B1'] = 'node_sequence'
    sheet['C1'] = 'time_sequence'
    sheet['D1'] = 'volume'
    sheet['E1'] = 'departure_time'
    sheet['F1'] = 'arrival_time'
    sheet['G1'] = 'total_travel_time'      
    for p in range(1, g_number_of_passengers):
        row = p + 1
        volume = g_passenger_list[p].volume
        departure_time = g_passenger_list[p].departure_time
        arrival_time = g_passenger_list[p].arrival_time
        total_travel_time = g_passenger_list[p].arrival_time - g_passenger_list[p].departure_time
        node_sequence = ";".join(str(node) for node in g_passenger_list[p].node_sequence_upper_bound)
        time_sequence = ";".join(str(time) for time in g_passenger_list[p].time_sequence_upper_bound)
        sheet.cell(row = row, column = 1, value = p)
        sheet.cell(row = row, column = 2, value = node_sequence)
        sheet.cell(row = row, column = 3, value = time_sequence)
        sheet.cell(row = row, column = 4, value = volume)
        sheet.cell(row = row, column = 5, value = departure_time)
        sheet.cell(row = row, column = 6, value = arrival_time)
        sheet.cell(row = row, column = 7, value = total_travel_time)           
    workbook.save('output_upper_bound.xlsx')

    # output upper bound and lower bound
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'output_gap'
    sheet['A1'] = 'Iteration'
    sheet['B1'] = 'lower_bound'
    sheet['C1'] = 'upper_bound'
    sheet['D1'] = 'Best_lower_bound'
    sheet['E1'] = 'Best_upper_bound'
    sheet['F1'] = 'Gap'

    for itr in range(1, maximum_iteration_step):
        row = itr + 1
        lower_bound_itr=lower_bound[itr]
        upper_bound_itr=upper_bound[itr]
        Best_lower_bound=max(lower_bound[:itr])
        Best_upper_bound=min(upper_bound[:itr])
        # gap=(Best_upper_bound-Best_lower_bound)/Best_upper_bound
        sheet.cell(row = row, column = 1, value = itr)
        sheet.cell(row = row, column = 2, value = lower_bound_itr)
        sheet.cell(row = row, column = 3, value = upper_bound_itr)
        sheet.cell(row = row, column = 4, value = Best_lower_bound)
        sheet.cell(row = row, column = 5, value = Best_upper_bound)
        # sheet.cell(row = row, column = 6, value = gap)

    workbook.save('output_gap.xlsx')

    return()

if __name__=='__main__':
    print('Reading data......') 
    # define parameter
    maximum_iteration_step = 2
    cap = 180
    headway = 1
    rou = 50
    # penalty = 10
    # input data  
    g_read_input_data()
    g_add_new_node()
    g_add_new_passenger_link()
    g_add_new_train_link()
    g_generate_in_out_going_link()
    g_generate_node_link_map()


    g_initialize_multiplier_and_cost()
    
    lower_bound = numpy.zeros([maximum_iteration_step + 1])
    lower_bound_1 = numpy.zeros([maximum_iteration_step + 1])
    lower_bound_2 = numpy.zeros([maximum_iteration_step + 1])
    lower_bound_3 = numpy.zeros([maximum_iteration_step + 1])

    upper_bound = numpy.zeros([maximum_iteration_step + 1])
    
    # time_dependent_link_state_for_passengers = time_dependent_initial_link_state
    for iteration_step in range(0, maximum_iteration_step):
        print(iteration_step)
        # initialize time dependent link volume    
        time_dependent_link_volume_for_passengers = numpy.zeros([g_number_of_passenger_links, g_number_of_time_intervals])
        time_dependent_link_volume_for_trains = numpy.zeros([g_number_of_train_links, g_number_of_time_intervals])
        # passenger and train assignment by dynamic programming 
        # g_time_dependent_dynamic_programming_for_trains()
        g_time_dependent_dynamic_programming_for_passengers()  
        
        g_update_cost()
        g_time_dependent_dynamic_programming_for_trains()
        g_generate_and_calculate_upper_bound()


        g_calcualte_lower_bound()
        # update multiplier
        step_size = 1 / (1 + iteration_step)
        g_update_multiplier()

    # output data
    g_write_output_data()

    f = open("output_volume.csv", "w")
    for l in range(g_number_of_passenger_links):
        for t in range(g_number_of_time_intervals):
            f.write(str(time_dependent_link_volume_for_passengers[l][t])+",")
        f.write("\n")
    f.close()

    

    
    